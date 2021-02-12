#! python3
# multiplicationTable.py - takes a number N from the command line and
# creates an N×N multiplication table in an Excel spreadsheet.

import openpyxl, sys
from openpyxl.styles import NamedStyle, Font

#store the N to variable named 'N'
if len(sys.argv) > 1:
	N = int(sys.argv[1])

#create a new workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

#style font to bold for header
default_bold = NamedStyle(name='default_bold')
default_bold.font = Font(bold=True)

#initiate the values in row '1' and column 'A' and set them to 'bold'
for row_num in range(2, N+2):
	sheet.cell(row = row_num, column = 1).value = row_num - 1
	sheet.cell(row = row_num, column = 1).style = default_bold

for column_num in range(2, N+2):
	sheet.cell(row = 1, column = column_num).value = column_num - 1
	sheet.cell(row = 1, column = column_num).style = default_bold

#fill up the N*N multiplication table
for column_num in range(2, N+2):
	for row_num in range(2, N+2):
		sheet.cell(row = row_num, column= column_num).value = sheet.cell(row = row_num, column= 1).value * sheet.cell(row = 1, column= column_num).value 

#save your worksheet as 'multiplicationTable.xlsx'
wb.save('F:\\multiplicationTable.xlsx')