#! python3

# blankRowInserter.py - takes two integers and a filename
# string as command line arguments.
# Starting at row N, the program should insert M blank rows
# into the spreadsheet.

import openpyxl, sys

# substitue user command argument to variable 'N', 'M' and 'filename'
if len(sys.argv) > 1:
	N = sys.argv[1]
	M = sys.argv[2]
	filename = sys.argv[3]

# load workbook using the 'filename'
wb = openpyxl.load_workbook('YOUR FILE PATH')
sheet = wb.active

# select the maximum column number to fill with 'blank space'
column_num = sheet.max_column

# Starting at row N, insert M blank rows
for column_num in range(1, column_num):
	for row_num in range(N, N+M):
		sheet.cell(row=row_num, column=column_num).value = ''

wb.save('YOUR SAVE FILE PATH')
