#! python3
# Spreadsheet Cell Inverter.py -  invert the row and column of the cells in the spreadsheet.

import openpyxl

# create two workbook 'wb' for original file and 'wb2' a blank workbook for storing inverted sheet
wb = openpyxl.load_workbook('F:\\python\\example.xlsx')
wb2 = openpyxl.Workbook()

sheet1 = wb.active
sheet2 = wb2.active

# store maximum column and maximum row numbers from 'sheet1'
column_maxnum = sheet1.max_column
row_maxnum = sheet1.max_row

# invert the row and column of the original 'sheet1' to blank 'sheet2'
for column_num in range(1, column_maxnum+1):
	for row_num in range(1, row_maxnum+1):
		sheet2.cell(row=column_num, column=row_num).value = sheet1.cell(row=row_num, column=column_num).value

# save the inverted worksheet
wb2.save('F:\\example-Copy.xlsx')